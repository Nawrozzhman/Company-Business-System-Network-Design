#!/usr/bin/env python3
"""
Scientific Competition Management System - Auto Generator
Generates complete Excel .xlsm workbook with VBA code
Author: Competition Management Team
Version: 1.0 PART 1
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter
import os
from datetime import datetime

class CompetitionSystemGenerator:
    """Generates Scientific Competition Management System workbook"""
    
    def __init__(self, output_filename="Scientific_Competition_Management_System.xlsm"):
        self.filename = output_filename
        self.wb = None
        self.ws_dashboard = None
        self.ws_students = None
        self.ws_settings = None
        self.ws_help = None
        
    def create_workbook(self):
        """Create new workbook"""
        print("[1/8] Creating workbook structure...")
        self.wb = openpyxl.Workbook()
        self.wb.remove(self.wb.active)  # Remove default sheet
        
    def create_sheets(self):
        """Create all required sheets"""
        print("[2/8] Creating worksheets...")
        
        # Dashboard
        self.ws_dashboard = self.wb.create_sheet("Dashboard", 0)
        
        # User Interface Sheets
        self.ws_students = self.wb.create_sheet("Students", 1)
        self.ws_settings = self.wb.create_sheet("Settings", 2)
        self.ws_help = self.wb.create_sheet("Help", 3)
        
        # Data Tables (hidden)
        ws_tbl_students = self.wb.create_sheet("tbl_Students", 4)
        ws_tbl_students.sheet_state = 'hidden'
        
        ws_tbl_settings = self.wb.create_sheet("tbl_Settings", 5)
        ws_tbl_settings.sheet_state = 'hidden'
        
        ws_grades = self.wb.create_sheet("tbl_Grades", 6)
        ws_grades.sheet_state = 'hidden'
        
        ws_sections = self.wb.create_sheet("tbl_Sections", 7)
        ws_sections.sheet_state = 'hidden'
        
        ws_validation = self.wb.create_sheet("_Validation_Lists", 8)
        ws_validation.sheet_state = 'hidden'
        
        ws_log = self.wb.create_sheet("_System_Log", 9)
        ws_log.sheet_state = 'hidden'
        
    def create_dashboard(self):
        """Create Dashboard UI"""
        print("[3/8] Building Dashboard...")
        
        ws = self.ws_dashboard
        
        # Set column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 40
        
        # Title
        ws['A1'] = "SCIENTIFIC COMPETITION MANAGEMENT SYSTEM"
        ws['A1'].font = Font(name='Calibri', size=16, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 30
        
        ws['A2'] = "Version 1.0 - Educational System"
        ws['A2'].font = Font(name='Calibri', size=10, italic=True)
        ws['A2'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A2:B2')
        ws.row_dimensions[2].height = 20
        
        # Main Menu
        row = 4
        ws[f'A{row}'] = "MAIN MENU"
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
        
        row = 6
        menu_items = [
            ("Students", "Manage student records"),
            ("Teachers & Judges", "Manage teacher/judge information"),
            ("Competitions", "Create and manage competitions"),
            ("Subjects", "Configure subjects"),
            ("Question Bank", "Manage questions"),
            ("Participants", "Select participants for competitions"),
            ("Score Entry", "Enter competition scores"),
            ("Results", "View competition results"),
            ("Rankings", "View rankings"),
            ("Reports", "Generate reports"),
            ("Certificates", "Generate certificates"),
            ("Settings", "System configuration"),
            ("Help", "User guide and documentation"),
        ]
        
        for item, description in menu_items:
            ws[f'A{row}'] = item
            ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
            ws[f'B{row}'] = description
            ws[f'B{row}'].font = Font(name='Calibri', size=10)
            ws.row_dimensions[row].height = 18
            row += 1
        
        # Statistics Section
        row += 2
        ws[f'A{row}'] = "QUICK STATISTICS"
        ws[f'A{row}'].font = Font(name='Calibri', size=12, bold=True)
        
        row += 2
        stats = [
            ("Total Students:", 0),
            ("Total Competitions:", 0),
            ("Active Competitions:", 0),
            ("Total Subjects:", 0),
        ]
        
        for stat, value in stats:
            ws[f'A{row}'] = stat
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'B{row}'] = value
            ws[f'B{row}'].font = Font(name='Calibri', size=10)
            row += 1
        
    def create_students_sheet(self):
        """Create Students Management UI"""
        print("[4/8] Building Student Management sheet...")
        
        ws = self.ws_students
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 30
        
        # Header
        ws['A1'] = "STUDENT MANAGEMENT"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:F1')
        ws.row_dimensions[1].height = 25
        
        # Action Buttons Row
        ws['A2'] = "[Dashboard] [Add] [Edit] [Delete] [Search] [Clear] [Print] [Refresh]"
        ws['A2'].font = Font(name='Calibri', size=9, italic=True)
        ws.merge_cells('A2:F2')
        
        # Student Entry Form
        row = 4
        form_fields = [
            "Student ID:",
            "Student Name:",
            "Father Name:",
            "Grandfather Name:",
            "Gender:",
            "Grade:",
            "Section:",
            "Roll Number:",
            "Date of Birth:",
            "Phone:",
            "Address:",
            "School:",
            "Academic Year:",
            "Status:",
            "Notes:",
        ]
        
        for field in form_fields:
            ws[f'A{row}'] = field
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'B{row}'].font = Font(name='Noto Naskh Arabic', size=10)  # Pashto support
            row += 1
        
        # Student List Header
        row += 2
        ws[f'A{row}'] = "STUDENT LIST"
        ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws.merge_cells(f'A{row}:F{row}')
        
        row += 1
        headers = ["StudentID", "Name", "Father Name", "Grade", "Section", "Status"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col)
            cell.value = header
            cell.font = Font(name='Calibri', size=10, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
        
    def create_settings_sheet(self):
        """Create Settings UI"""
        print("[5/8] Building Settings sheet...")
        
        ws = self.ws_settings
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 40
        
        # Header
        ws['A1'] = "SYSTEM SETTINGS"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.merge_cells('A1:B1')
        ws.row_dimensions[1].height = 25
        
        # Settings Form
        row = 3
        settings_fields = [
            "School Name:",
            "Principal Name:",
            "Academic Year:",
            "Competition Coordinator:",
            "Phone:",
            "Email:",
            "Address:",
            "Default Max Marks:",
            "Passing Percentage:",
        ]
        
        for field in settings_fields:
            ws[f'A{row}'] = field
            ws[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
            ws[f'B{row}'].font = Font(name='Calibri', size=10)
            row += 1
        
        # Save Button Info
        row += 2
        ws[f'A{row}'] = "[Save Settings] [Reset] [Backup] [Restore]"
        ws[f'A{row}'].font = Font(name='Calibri', size=9, italic=True)
        
    def create_help_sheet(self):
        """Create Help/User Guide"""
        print("[6/8] Building Help sheet...")
        
        ws = self.ws_help
        ws.column_dimensions['A'].width = 100
        
        # Header
        ws['A1'] = "HELP & USER GUIDE"
        ws['A1'].font = Font(name='Calibri', size=14, bold=True, color="FFFFFF")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws.row_dimensions[1].height = 25
        
        row = 3
        help_sections = [
            ("1. HOW TO ADD STUDENTS", [
                "- Click 'Add' button on Students sheet",
                "- Fill in all required fields",
                "- Student names can be in English or Pashto",
                "- Click Save to add student",
                "- Student ID is auto-generated",
            ]),
            ("2. HOW TO CREATE A COMPETITION", [
                "- Go to Competitions sheet",
                "- Click 'Add Competition' button",
                "- Enter competition details",
                "- Select participating grade/section",
                "- Click Save",
            ]),
            ("3. HOW TO ENTER SCORES", [
                "- Go to Score Entry sheet",
                "- Select competition",
                "- Select student",
                "- Enter marks obtained",
                "- System calculates automatically",
                "- Ranking updates in real-time",
            ]),
            ("4. PASHTO LANGUAGE SUPPORT", [
                "- All text fields support Pashto Unicode",
                "- Enter Pashto names directly",
                "- No conversion needed",
                "- Data is preserved when saved",
            ]),
            ("5. PRINTING & REPORTS", [
                "- Go to Reports sheet",
                "- Select report type",
                "- Click Preview or Print",
                "- Reports are formatted for A4 paper",
            ]),
            ("6. DATA BACKUP", [
                "- Click 'Backup' button in Settings",
                "- Choose save location",
                "- File is automatically backed up",
                "- Use 'Restore' to recover data",
            ]),
        ]
        
        for section_title, section_content in help_sections:
            ws[f'A{row}'] = section_title
            ws[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
            row += 1
            
            for line in section_content:
                ws[f'A{row}'] = line
                ws[f'A{row}'].font = Font(name='Calibri', size=10)
                row += 1
            
            row += 1
        
    def create_data_tables(self):
        """Create data tables in hidden sheets"""
        print("[7/8] Creating data tables...")
        
        # Get hidden sheets
        ws_students = self.wb["tbl_Students"]
        ws_settings = self.wb["tbl_Settings"]
        ws_grades = self.wb["tbl_Grades"]
        ws_sections = self.wb["tbl_Sections"]
        ws_validation = self.wb["_Validation_Lists"]
        
        # Students Table
        headers_students = [
            "StudentID", "StudentName", "FatherName", "GrandFatherName",
            "Gender", "Grade", "Section", "RollNumber", "DateOfBirth",
            "Phone", "Address", "School", "AcademicYear", "Status",
            "CreatedDate", "Notes"
        ]
        
        for col, header in enumerate(headers_students, 1):
            ws_students.cell(row=1, column=col, value=header)
            ws_students.cell(row=1, column=col).font = Font(bold=True)
        
        # Settings Table
        settings_data = [
            ["SchoolName", "School Name", "Text", "Name of the school"],
            ["PrincipalName", "Principal", "Text", "Name of principal"],
            ["AcademicYear", "2024-2025", "Text", "Current academic year"],
            ["CompetitionCoordinator", "Coordinator", "Text", "Competition coordinator"],
            ["DefaultMaxMarks", "100", "Number", "Default maximum marks"],
            ["PassingPercentage", "40", "Number", "Passing percentage"],
            ["Phone", "", "Text", "School contact number"],
            ["Email", "", "Text", "School email"],
            ["Address", "", "Text", "School address"],
        ]
        
        ws_settings.cell(row=1, column=1, value="SettingKey")
        ws_settings.cell(row=1, column=2, value="SettingValue")
        ws_settings.cell(row=1, column=3, value="DataType")
        ws_settings.cell(row=1, column=4, value="Description")
        
        for row, data in enumerate(settings_data, 2):
            for col, value in enumerate(data, 1):
                ws_settings.cell(row=row, column=col, value=value)
        
        # Grades Table
        ws_grades.cell(row=1, column=1, value="GradeID")
        ws_grades.cell(row=1, column=2, value="GradeName")
        ws_grades.cell(row=1, column=3, value="Description")
        
        for i in range(1, 11):
            ws_grades.cell(row=i+1, column=1, value=i)
            ws_grades.cell(row=i+1, column=2, value=str(i))
            ws_grades.cell(row=i+1, column=3, value=f"Grade {i}")
        
        # Sections Table
        sections = ["A", "B", "C", "D", "E"]
        ws_sections.cell(row=1, column=1, value="SectionID")
        ws_sections.cell(row=1, column=2, value="SectionName")
        ws_sections.cell(row=1, column=3, value="Description")
        
        for idx, section in enumerate(sections, 1):
            ws_sections.cell(row=idx+1, column=1, value=idx)
            ws_sections.cell(row=idx+1, column=2, value=section)
            ws_sections.cell(row=idx+1, column=3, value=f"Section {section}")
        
        # Validation Lists
        ws_validation.cell(row=1, column=1, value="Gender")
        ws_validation.cell(row=2, column=1, value="M")
        ws_validation.cell(row=3, column=1, value="F")
        
        ws_validation.cell(row=1, column=3, value="Status")
        ws_validation.cell(row=2, column=3, value="Active")
        ws_validation.cell(row=3, column=3, value="Inactive")
        
        # System Log
        ws_log = self.wb["_System_Log"]
        ws_log.cell(row=1, column=1, value="Timestamp")
        ws_log.cell(row=1, column=2, value="Action")
        
    def save_workbook(self):
        """Save the workbook"""
        print(f"\n✓ Saving workbook as {self.filename}...")
        self.wb.save(self.filename)
        print(f"✓ Workbook saved successfully!")
        
    def generate(self):
        """Generate complete workbook"""
        print("\n" + "="*70)
        print("SCIENTIFIC COMPETITION MANAGEMENT SYSTEM - PART 1")
        print("Auto-Generator v1.0")
        print("="*70 + "\n")
        
        self.create_workbook()
        self.create_sheets()
        self.create_dashboard()
        self.create_students_sheet()
        self.create_settings_sheet()
        self.create_help_sheet()
        self.create_data_tables()
        
        self.save_workbook()
        
        print("\n" + "="*70)
        print("✓ WORKBOOK GENERATION COMPLETE!")
        print("="*70)
        print(f"\n✓ File created: {self.filename}")
        print("\nNEXT STEPS:")
        print("1. Open the generated Excel file")
        print("2. Review the Dashboard, Students, Settings, and Help sheets")
        print("3. Add VBA code using the VBA_CODE.txt file")
        print("4. Run the InitializeWorkbook() macro")
        print("5. System will be ready to use!")
        print("\n" + "="*70 + "\n")


def main():
    """Main entry point"""
    try:
        generator = CompetitionSystemGenerator()
        generator.generate()
        
        print("\nIMPORTANT NOTES:")
        print("-" * 70)
        print("✓ Excel workbook has been created successfully")
        print("✓ All sheets and tables are configured")
        print("✓ Check the VBA_CODE.txt file for Python/VBA modules")
        print("✓ Follow VBA_INSTALLATION_GUIDE.txt for detailed instructions")
        print("-" * 70 + "\n")
        
        return True
        
    except Exception as e:
        print(f"\n✗ Error during generation: {str(e)}")
        print("Please check the error message above and try again.")
        return False


if __name__ == "__main__":
    success = main()
    if success:
        print("SUCCESS! Your Excel file is ready to use.")
    else:
        print("FAILED! Please review the errors above.")
